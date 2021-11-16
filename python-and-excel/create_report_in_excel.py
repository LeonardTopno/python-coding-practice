#pip3 install openpyxl
import datetime
      
#Reading from Excelsheet
from openpyxl import load_workbook 

wb1 = load_workbook("student_games.xlsx") #"/Users/leo/MyWork/python-coding-practice/student_gmaes.xlsx"

sheet = wb1.active

today = datetime.date.today()

print("Today's date:", today)

student_born_in_May = []
student_not_born_in_Dec = []
students_admitted_in_7_days = []
for row_ in range(2, sheet.max_row+1):
    
    cell_obj = sheet.cell(row=row_, column=7)
    #print(cell_obj.value)
    
    if cell_obj.value == 5:
        student_name = sheet.cell(row=row_, column=2) 
        student_born_in_May.append(student_name.value)
    
    if cell_obj.value != 12:
        student_name = sheet.cell(row=row_, column=2) 
        student_not_born_in_Dec.append(student_name.value)

    cell_obj_adm = sheet.cell(row=row_, column=3)
    #print(cell_obj_adm.value.date()) 
    
    no_of_days_passed = (today-cell_obj_adm.value.date()).days
    #print("Print Num of Days:", no_of_days_passed)

    if no_of_days_passed <=7:
        student_name = sheet.cell(row=row_, column=2)
        students_admitted_in_7_days.append(student_name.value)

    



print("Student Born In May", student_born_in_May)
print("Student Not Born In Dec", student_not_born_in_Dec)
print("Student Admitted in 7 Days", students_admitted_in_7_days)
        
    
#Writing
from openpyxl import Workbook

workbook = Workbook()

sheet = workbook.active
#sheet_1 = workbook.create_sheet("Sheet1")

sheet_2 = workbook.create_sheet("Sheet2")
sheet_3 = workbook.create_sheet("Sheet3")


sheet["A1"] = "May_Month_Birthday"
sheet_2["A1"] = "Last_7_days_Admission"
sheet_3["A1"] = "Birthday_Not_In_December" 

sheet.freeze_panes = "A1"  

for counter, elem in enumerate(student_born_in_May, start=2):
    sheet.cell(row=counter, column=1, value=elem)

for counter, elem in enumerate(students_admitted_in_7_days, start=2):
    sheet_2.cell(row=counter, column=1, value=elem)

for counter, elem in enumerate(student_not_born_in_Dec, start=2):
    sheet_3.cell(row=counter, column=1, value=elem)


workbook.save(filename="report.xlsx")
print("Report Created Successfully")
